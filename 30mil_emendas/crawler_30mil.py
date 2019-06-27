import pymongo
import openpyxl
import requests


def operacoes_mongodb():
	'''
		Operações relacionadas ao banco, inserções.
	'''

	myclient = pymongo.MongoClient("mongodb://localhost:27017/")

	mydb = myclient["Teste"]
	#print(myclient.list_database_names())

	mycol = mydb["colteste"]

	print(mycol.find_one())



def operacoes_obtencao_arquivos():
	'''
		Operações de coleta de arquivos.
	'''

	workbook = openpyxl.load_workbook("teste.xlsx")
	worksheet = workbook.active

	coluna_com_id = 1
	coluna_com_nome = 2
	coluna_com_link = 3	
	for row1 in range(2, worksheet.max_row + 1): #ignorando o cabeçalho
		ID = worksheet.cell(row = row1, column = coluna_com_id).value
		nome = worksheet.cell(row = row1, column = coluna_com_nome).value
		link = worksheet.cell(row = row1, column = coluna_com_link).hyperlink.target
		#texto = worksheet.cell(row = row1, column = coluna_com_link).value
		print(ID)
		print(nome)
		print(link)
		requisicao = requests.get(link, allow_redirects = True)
		with open('teores_emendas/' + nome + '.pdf', 'wb') as arquivo:
			arquivo.write(requisicao.content)




def main():
	operacoes_obtencao_arquivos()



if __name__ == "__main__":
    main()
