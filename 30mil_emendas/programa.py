import pymongo
import pandas as pd
import openpyxl
import requests

def configuracoes():
	pd.set_option('display.max_rows', 500)
	pd.set_option('display.max_columns', 500)


def operacoes_mongodb():
	myclient = pymongo.MongoClient("mongodb://localhost:27017/")

	mydb = myclient["Teste"]
	#print(myclient.list_database_names())

	mycol = mydb["colteste"]

	print(mycol.find_one())



def operacoes_obtencao_arquivos():
	'''print()
	#workbook_dataframe = pd.read_excel(, skip_footer = (rows - nrows - 1))

	workbook = pd.ExcelFile("emendas/emendas.xlsx")

	# get the total number of rows (assuming you're dealing with the first sheet)
	rows = workbook.book.sheet_by_index(0).nrows

	# define how many rows to read
	nrows = 10

	# subtract the number of rows to read from the total number of rows (and another 1 for the header)
	workbook_dataframe = pd.read_excel(workbook, skipfooter = (rows - nrows - 1))

	print(workbook_dataframe)
	#workbook_dataframe
	'''

	'''
	workbook = openpyxl.load_workbook("teste.xlsx")
	worksheet = workbook.active
	#print(worksheet['A1'].value)
	#print(worksheet.cell(row=1, column=1).value)
	#exit(0)
	for row1 in range(1, worksheet.max_row+1):
		text = worksheet.cell(row = row1, column = 1).value
		print(text)
	'''

	workbook = openpyxl.load_workbook("teste.xlsx")
	worksheet = workbook.active

	coluna_com_id = 1
	coluna_com_nome = 2
	coluna_com_link = 3	
	for row1 in range(2, worksheet.max_row + 1): #ignorando o cabeçalho
		nome = worksheet.cell(row = row1, column = coluna_com_nome).value
		link = worksheet.cell(row = row1, column = coluna_com_link).hyperlink.target
		#texto = worksheet.cell(row = row1, column = coluna_com_link).value
		print(nome)
		print(link)
		requisicao = requests.get(link, allow_redirects = True)
		with open('teores_emendas/' + nome + '.pdf', 'wb') as arquivo:
			arquivo.write(requisicao.content)


	#column_indices = [5]
	'''
	for row in range(3, 5): #worksheet.max_row+1): 
		text = worksheet.cell(1, row = row)
		print(text)
		for col in column_indices:
			filelocation = worksheet.cell(column=col, row=row ) #this is hyperlink
			text = worksheet.cell(column=col+1, row=row) # thi is your text 
			worksheet.cell(column=col+1,row=row).value = '=HYPERLINK("'+filelocation.value+'","'+text.value+'")'
	'''


	#workbook.save('teste.xlsx')



def main():
	#configuracoes()
	operacoes_obtencao_arquivos()



if __name__ == "__main__":
    main()
